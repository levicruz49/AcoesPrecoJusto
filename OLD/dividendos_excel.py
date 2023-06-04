import pandas as pd
import yfinance as yf
import asyncio
import pytz
import openpyxl
from openpyxl.utils import column_index_from_string

yf.pdr_override()

async def get_dividends_yfinance(ticker, period):
    stock = yf.Ticker(ticker)
    if hasattr(stock, 'dividends'):
        history = stock.history(period='max')
        if 'Dividends' in history.columns:
            dividends = history['Dividends']
        else:
            dividends = pd.Series(dtype='float64')

        if isinstance(dividends.index, pd.DatetimeIndex):
            dividends.index = dividends.index.tz_convert('America/Sao_Paulo')

        now = pd.Timestamp.now(tz=pytz.timezone('America/Sao_Paulo'))

        if period == '1y':
            dividends = dividends[dividends.index > (now - pd.DateOffset(years=1)).to_pydatetime()]
        elif period == '6y':
            dividends = dividends[dividends.index > (now - pd.DateOffset(years=6)).to_pydatetime()]
        elif period == 'max':
            pass
        else:
            return None

        if len(dividends) > 0:
            return round(dividends.sum(), 2)
        else:
            return None
    else:
        return None

async def main(tickers):
    dividend_data = {}
    for ticker in tickers:
        try:
            dividend_data[ticker] = {
                '1y': await get_dividends_yfinance(ticker, '1y'),
                '6y': await get_dividends_yfinance(ticker, '6y'),
                'max': await get_dividends_yfinance(ticker, 'max')
            }
        except Exception as e:
            print(f"Erro ao buscar dados para {ticker}: {e}")
            continue

    return dividend_data

def update_spreadsheet(dividend_data, workbook, start_row):
    sheet = workbook["Acoes"]
    for i, (ticker, data) in enumerate(dividend_data.items(), start=start_row):
        ticker = ticker.replace('.SA', '')  # Remover o sufixo '.SA'
        sheet.cell(row=i + 1, column=column_index_from_string('E'), value=data['1y'])
        sheet.cell(row=i + 1, column=column_index_from_string('F'), value=data['6y'])
        sheet.cell(row=i + 1, column=column_index_from_string('G'), value=data['max'])
    workbook.save("C:\\Users\\mrcr\\Desktop\\preco_justo.xlsx")


if __name__ == "__main__":
    workbook = openpyxl.load_workbook("C:\\Users\\mrcr\\Desktop\\preco_justo.xlsx", keep_vba=True)
    sheet = workbook["Acoes"]

    # Obter lista de tickers da coluna C, começando na linha 3
    tickers = [cell.value for cell in sheet['C'][2:] if isinstance(cell.value, str)]
    tickers_with_suffix = [ticker + ".SA" for ticker in tickers]

    dividend_data = asyncio.run(main(tickers_with_suffix))

    update_spreadsheet(dividend_data, workbook, 2)


